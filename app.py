import io
import re
import unicodedata
from urllib.parse import urlparse, parse_qs

import pandas as pd
import requests
import streamlit as st
from rapidfuzz import fuzz


# ============================================================
# CONFIG
# ============================================================

DEFAULT_COA_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "17o01929ZP5jkHupJkOU5XhnPU72HrOfx/edit?gid=461974411#gid=461974411"
)

# Akun-akun yang scope-nya memang ditangani tool ini. Dipakai untuk
# menampilkan daftar ke pengguna dan untuk memvalidasi file yang diupload,
# supaya ketahuan dari awal kalau ada file dari akun di luar cakupan.
ALLOWED_ACCOUNTS = {
    "54100000": "ATK, Foto Copy dan Cetakan",
    "54200000": "Telepon",
    "54300000": "Biaya Komputer/IT",
    "54400000": "BBM/Transport",
    "54500000": "Biaya Transport Lainnya",
    "54600000": "Listrik & Air",
    "54700000": "Sewa",
    "54800000": "Biaya Elektronik",
    "54900000": "Biaya Perlengkapan Kantor",
    "55200000": "Biaya Pengiriman",
    "55300000": "Biaya Konsumsi",
    "55900000": "Biaya Survei Regional",
    "55999000": "Biaya Kantor Lainnya",
    "56100000": "Perawatan Gedung/Kantor",
    "56200000": "Perawatan Kendaraan",
    "56300000": "Perawatan Komputer/IT",
    "56400000": "Perawatan Elektronik",
    "56500000": "Perawatan Perlengkapan Kantor/Furniture",
    "56600000": "Perawatan Software",
    "57200000": "Rugi Penghapusan Aset",
    "59910000": "Biaya Kerugian Pinjaman",
    "59921000": "Biaya Administrasi Bank",
    "59922000": "Sumbangan",
    "59923000": "Iuran Keamanan",
    "59924000": "Iuran Kebersihan",
}

RESULT_COLUMNS = [
    "SOURCE FILE",
    "KODE AKUN",
    "NAMA AKUN",
    "VOUCHER NO.",
    "TRANS. DATE",
    "ENTRY DATE",
    "DESCRIPTION",
    "DEBIT BASE",
    "CREDIT BASE",
    "DEBIT FOREX",
    "CREDIT FOREX",
    "DOCUMENT NO.",
    "COA CRITERIA",
    "STATUS",
    "MATCH SCORE",
    "MATCHED CRITERIA",
    "ALTERNATIVE COA",
    "ALTERNATIVE ACCOUNT",
    "ALTERNATIVE FILE",
    "REVIEW REASON",
]


# ============================================================
# TEXT / NORMALIZATION
# ============================================================

STOPWORDS = {
    "dan", "atau", "yang", "untuk", "dari", "ke", "di", "pada", "dengan",
    "serta", "dll", "lainnya", "lain", "kebutuhan", "pembelian", "pembelian",
    "biaya", "dibeli", "dibayarkan", "pengadaan", "pengeluaran", "rutin",
    "kantor", "cabang", "unit", "masa", "manfaat", "lebih", "tahun", "harga",
    "mulai", "item", "pcs", "buah", "dus", "pack", "pak", "lembar", "buahnya",
    # Jabatan / kata pengantar nama yang sering muncul di deskripsi voucher
    # dan tidak boleh ikut menentukan kecocokan akun.
    "bm", "asmen", "fsa", "spv", "an", "staf", "staff", "karyawan",
    "mingguan", "harian", "bulanan",
    # Kata kerja generik yang muncul di berbagai kategori berbeda ("isi
    # cutter" di ATK vs "isi ulang galon" di Konsumsi). Kalau dianggap
    # token bermakna, kata ini jadi jembatan salah antar akun yang tidak
    # berhubungan sama sekali.
    "isi", "ulang", "beli",
}

ALIASES = {
    "tissu": "tisu",
    "tissue": "tisu",
    "fotocopy": "fotokopi",
    "foto copy": "fotokopi",
    "foto kopi": "fotokopi",
    "bolpoint": "pulpen",
    "ballpoint": "pulpen",
    "post-it": "post it",
    "postit": "post it",
    "posit": "post it",
    "sticky note": "post it",
    "kertas f 4": "kertas f4",
    "kertas folio": "kertas",
    "formulir": "form",
    "formulir pengajuan": "form",
    "dok": "dokumen",
    "arsip": "dokumen",
    "aqua": "air minum",
    "air mineral": "air minum",
    "17 agustus": "hut ri",
    "tujuh belas agustus": "hut ri",
    "kemerdekaan ri": "hut ri",
}

# Frasa pada deskripsi yang menandakan kegiatan sosial kemasyarakatan resmi
# (RT/RW, desa, kelurahan, HUT RI, dsb). Dipakai untuk mencocokkan ke akun
# yang Kriteria Transaksinya menyebut "sosial kemasyarakatan" / "pemerintahan
# resmi" / "bencana alam" tanpa harus hardcode ke kode akun tertentu.
GOV_COMMUNITY_HINTS = {
    "hut ri", "partisipasi kegiatan", "kegiatan sosial", "kegiatan warga",
    "kemerdekaan",
}

# Kata yang menandakan ORMAS dan harus MENGECUALIKAN transaksi dari kriteria
# "pemerintahan resmi (bukan ORMAS)", sesuai aturan bisnis.
ORMAS_HINTS = {
    "ormas", "organisasi", "karang taruna", "paguyuban", "komunitas",
}

# Tokens that are too generic to trigger an alternative account by themselves.
AMBIGUOUS_TOKENS = {
    "air", "uang", "kantor", "anggota", "pengajuan", "pengeluaran",
    "pembayaran", "top", "up", "transaksi", "kegiatan", "operasional",
    "rutin", "bulan", "tahun", "baru", "lama", "khusus", "rincian",
    "form", "dokumen", "lain", "lainnya",
}


def normalize_text(value):
    if value is None:
        return ""

    value = str(value)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower().replace("\xa0", " ")

    # Common Indonesian transaction-description variants.
    for source, target in sorted(ALIASES.items(), key=lambda x: -len(x[0])):
        value = re.sub(rf"\b{re.escape(source)}\b", target, value)

    value = re.sub(r"[^a-z0-9]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def meaningful_tokens(text):
    return {
        token
        for token in normalize_text(text).split()
        if len(token) >= 3 and token not in STOPWORDS
    }


def split_criteria(criteria):
    """
    Converts a long COA criteria cell into practical matching phrases.

    Parenthetical text is kept separately because it often contains
    accounting conditions such as useful-life / minimum-value rules.
    """
    criteria = "" if criteria is None else str(criteria)
    if not criteria.strip():
        return []

    parts = []

    # Keep the complete criteria as a low-priority phrase.
    whole = normalize_text(criteria)
    if whole:
        parts.append(whole)

    # Remove parenthetical conditions for item/keyword extraction.
    without_parentheses = re.sub(r"\([^)]*\)", " ", criteria)

    # Most of the COA uses comma / semicolon / slash separated examples.
    chunks = re.split(
        r"[,;/:\n]+|\s+\b(?:dan|atau)\b\s+",
        without_parentheses,
        flags=re.IGNORECASE,
    )

    for chunk in chunks:
        chunk = normalize_text(chunk)
        if not chunk:
            continue

        tokens = meaningful_tokens(chunk)
        if not tokens:
            continue

        # Ignore very generic fragments.
        if len(tokens) == 1 and next(iter(tokens)) in {
            "kegiatan", "operasional", "transaksi", "penggunaan"
        }:
            continue

        parts.append(chunk)

    # De-duplicate while preserving order.
    result = []
    seen = set()
    for item in parts:
        if item not in seen:
            seen.add(item)
            result.append(item)

    return result


def extract_amount(text):
    """
    Extracts a simple rupiah amount from a description.
    Used only as supporting evidence, never as the sole match reason.
    """
    if not text:
        return None

    patterns = [
        r"(?:rp\.?|idr)\s*([0-9][0-9\.,]*)",
        r"([0-9][0-9\.,]*)\s*(?:ribu|rb|juta)",
    ]

    for pattern in patterns:
        match = re.search(pattern, str(text), flags=re.IGNORECASE)
        if match:
            raw = match.group(1).replace(".", "").replace(",", "")
            try:
                return float(raw)
            except ValueError:
                pass

    return None


# ============================================================
# GOOGLE SHEETS COA
# ============================================================

def parse_google_sheet_url(url):
    """
    Accepts a normal Google Sheets edit URL and extracts spreadsheet ID + gid.
    """
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not match:
        raise ValueError("Spreadsheet ID tidak ditemukan dari URL Google Sheets.")

    spreadsheet_id = match.group(1)

    parsed = urlparse(url)
    query = parse_qs(parsed.query)

    gid = query.get("gid", [None])[0]
    if not gid and parsed.fragment:
        fragment_query = parse_qs(parsed.fragment)
        gid = fragment_query.get("gid", [None])[0]

    if not gid:
        gid = "0"

    return spreadsheet_id, gid


@st.cache_data(ttl=300, show_spinner=False)
def load_coa_from_google(url):
    spreadsheet_id, gid = parse_google_sheet_url(url)

    gviz_url = (
        f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
        f"/gviz/tq?tqx=out:csv&gid={gid}"
    )

    response = requests.get(
        gviz_url,
        timeout=30,
        headers={"User-Agent": "Mozilla/5.0"},
    )
    response.raise_for_status()

    content_type = response.headers.get("content-type", "").lower()
    if "text/html" in content_type and "csv" not in content_type:
        # Google may return an HTML permission page if the sheet is not public.
        preview = response.text[:500].lower()
        if "sign in" in preview or "request access" in preview:
            raise PermissionError(
                "Google Sheets tidak dapat dibaca. Pastikan file "
                "dapat diakses dengan mode 'Anyone with the link / Siapa saja "
                "yang memiliki link dapat melihat'."
            )

    df = pd.read_csv(io.StringIO(response.text), dtype=str).fillna("")

    # Standardize column names.
    df.columns = [normalize_text(c).upper() for c in df.columns]

    # Expected columns are A=KODE AKUN, B=NAMA AKUN, C=KRITERIA TRANSAKSI.
    column_map = {}
    for col in df.columns:
        n = normalize_text(col)
        if "kode akun" in n or n in {"akun", "account", "account no"}:
            column_map[col] = "KODE AKUN"
        elif "nama akun" in n or n in {"account name", "nama"}:
            column_map[col] = "NAMA AKUN"
        elif "kriteria transaksi" in n or "criteria" in n:
            column_map[col] = "KRITERIA TRANSAKSI"
        elif "kelompok" in n or "group" in n:
            column_map[col] = "KELOMPOK"

    df = df.rename(columns=column_map)

    # If the header was not recognized, fall back to the first four columns.
    if "KODE AKUN" not in df.columns and len(df.columns) >= 1:
        df = df.rename(columns={df.columns[0]: "KODE AKUN"})
    if "NAMA AKUN" not in df.columns and len(df.columns) >= 2:
        remaining = [c for c in df.columns if c != "KODE AKUN"]
        if remaining:
            df = df.rename(columns={remaining[0]: "NAMA AKUN"})
    if "KRITERIA TRANSAKSI" not in df.columns and len(df.columns) >= 3:
        remaining = [c for c in df.columns if c not in {"KODE AKUN", "NAMA AKUN"}]
        if remaining:
            df = df.rename(columns={remaining[0]: "KRITERIA TRANSAKSI"})
    if "KELOMPOK" not in df.columns and len(df.columns) >= 4:
        remaining = [
            c for c in df.columns
            if c not in {"KODE AKUN", "NAMA AKUN", "KRITERIA TRANSAKSI"}
        ]
        if remaining:
            df = df.rename(columns={remaining[0]: "KELOMPOK"})

    for col in ["KODE AKUN", "NAMA AKUN", "KRITERIA TRANSAKSI", "KELOMPOK"]:
        if col not in df.columns:
            df[col] = ""

    df["KODE AKUN"] = df["KODE AKUN"].apply(clean_account_code)
    df["NAMA AKUN"] = df["NAMA AKUN"].astype(str).str.strip()
    df["KRITERIA TRANSAKSI"] = df["KRITERIA TRANSAKSI"].astype(str).str.strip()
    df["KELOMPOK"] = df["KELOMPOK"].astype(str).str.strip()

    df = df[df["KODE AKUN"] != ""].copy()
    df = df.drop_duplicates(subset=["KODE AKUN"], keep="first")

    return df.reset_index(drop=True)


# ============================================================
# XLS / HTML REPORT PARSER
# ============================================================

def clean_account_code(value):
    if value is None:
        return ""

    text = str(value).strip()
    digits = re.sub(r"\D", "", text)

    # Account code in the COA is generally an 8-digit code.
    if len(digits) >= 8:
        return digits[:8]

    return digits


def parse_account_metadata(text):
    text = str(text).replace("\xa0", " ")

    account_match = re.search(
        r"Account No\.\s*:\s*([0-9]+)",
        text,
        flags=re.IGNORECASE,
    )

    name_match = re.search(
        r"Account Name\s*:\s*(.*?)(?:,\s*Beginning Balance|$)",
        text,
        flags=re.IGNORECASE,
    )

    return (
        clean_account_code(account_match.group(1)) if account_match else "",
        name_match.group(1).strip() if name_match else "",
    )


def is_transaction_row(row):
    if len(row) < 5:
        return False

    voucher = str(row.iloc[0]).strip()
    trans_date = str(row.iloc[1]).strip()
    description = str(row.iloc[3]).strip()

    if not voucher or voucher.lower() == "nan":
        return False

    if not description or description.lower() == "nan":
        return False

    if re.search(r"TOTAL|ENDING BALANCE|GRAND TOTAL|ACCOUNT NO\.", voucher, re.I):
        return False

    # Transaction date must look like a date.
    parsed = pd.to_datetime(trans_date, dayfirst=True, errors="coerce")
    return pd.notna(parsed)


def parse_uploaded_xls(uploaded_file):
    """
    ERP reports are frequently HTML files saved with .xls extension.
    A single MDIS export can (and normally does) contain many accounts one
    after another, separated by "Account No. / Account Name" header rows.
    Valid NO DATA reports are returned as an empty dataframe.
    """
    raw = uploaded_file.getvalue()
    filename = uploaded_file.name
    raw_lower = raw.decode("utf-8", errors="ignore").lower()

    if "no data !!!" in raw_lower:
        return pd.DataFrame(columns=[
            "KODE AKUN", "NAMA AKUN", "VOUCHER NO.", "TRANS. DATE",
            "ENTRY DATE", "DESCRIPTION", "DEBIT BASE", "CREDIT BASE",
            "DEBIT FOREX", "CREDIT FOREX", "DOCUMENT NO.", "SOURCE FILE"
        ])

    tables = []
    try:
        tables = pd.read_html(io.BytesIO(raw), flavor="lxml")
    except Exception:
        try:
            tables = pd.read_html(io.BytesIO(raw))
        except Exception:
            tables = []

    if not tables:
        try:
            df_excel = pd.read_excel(
                io.BytesIO(raw),
                header=None,
                engine="xlrd" if filename.lower().endswith(".xls") else None,
            )
            tables = [df_excel]
        except Exception as exc:
            raise ValueError(f"Format file tidak dapat dibaca: {exc}")

    records = []
    current_code = ""
    current_name = ""

    for df in tables:
        if df.empty:
            continue
        df = df.fillna("")

        for _, row in df.iterrows():
            vals = row.astype(str).tolist()
            joined = " ".join(vals)

            if "Account No." in joined:
                code, name = parse_account_metadata(joined)
                if code:
                    current_code = code
                if name:
                    current_name = name
                continue

            if is_transaction_row(row):
                # Expected export layout:
                # 0 voucher, 1 trans date, 2 entry date, 3 description,
                # 4 base debit, 5 base credit,
                # 6 forex debit, 7 forex credit, 8 document no.
                records.append({
                    "KODE AKUN": current_code,
                    "NAMA AKUN": current_name,
                    "VOUCHER NO.": str(row.iloc[0]).strip(),
                    "TRANS. DATE": str(row.iloc[1]).strip(),
                    "ENTRY DATE": str(row.iloc[2]).strip(),
                    "DESCRIPTION": str(row.iloc[3]).strip(),
                    "DEBIT BASE": to_number(row.iloc[4]),
                    "CREDIT BASE": to_number(row.iloc[5]),
                    "DEBIT FOREX": to_number(row.iloc[6]) if len(row) > 6 else 0.0,
                    "CREDIT FOREX": to_number(row.iloc[7]) if len(row) > 7 else 0.0,
                    "DOCUMENT NO.": (
                        "" if len(row) <= 8 or pd.isna(row.iloc[8])
                        else str(row.iloc[8]).strip()
                    ),
                    "SOURCE FILE": filename,
                })

    return pd.DataFrame(records, columns=[
        "KODE AKUN", "NAMA AKUN", "VOUCHER NO.", "TRANS. DATE",
        "ENTRY DATE", "DESCRIPTION", "DEBIT BASE", "CREDIT BASE",
        "DEBIT FOREX", "CREDIT FOREX", "DOCUMENT NO.", "SOURCE FILE"
    ])


def to_number(value):
    if pd.isna(value):
        return 0.0

    text = str(value).strip().replace(",", "")
    if text in {"", "nan", "None"}:
        return 0.0

    try:
        return float(text)
    except ValueError:
        # Handle formatted values such as 1.234.567.
        text = re.sub(r"[^0-9.-]", "", text)
        try:
            return float(text) if text else 0.0
        except ValueError:
            return 0.0


# ============================================================
# MATCHING ENGINE
# ============================================================

def score_phrase(description, phrase):
    desc = normalize_text(description)
    phrase = normalize_text(phrase)

    if not desc or not phrase:
        return 0.0

    # Exact phrase is strongest.
    if re.search(rf"\b{re.escape(phrase)}\b", desc):
        return 1.0

    desc_tokens = meaningful_tokens(desc)
    phrase_tokens = meaningful_tokens(phrase)

    if not phrase_tokens:
        return 0.0

    intersection = desc_tokens & phrase_tokens

    # A single ambiguous token must never create a classification.
    if len(phrase_tokens) == 1:
        token = next(iter(phrase_tokens))
        if token in AMBIGUOUS_TOKENS:
            return 0.0
        if token in intersection and len(token) >= 4:
            return 0.88

    if intersection:
        overlap = len(intersection) / len(phrase_tokens)
        if overlap >= 1.0:
            return 0.95
        if overlap >= 0.75:
            return 0.88
        if overlap >= 0.50:
            return 0.72

    # Fuzzy matching only for phrases of meaningful length.
    if len(phrase) >= 6:
        fuzzy = fuzz.token_set_ratio(desc, phrase) / 100.0
        if fuzzy >= 0.88:
            return fuzzy * 0.85

    return 0.0


def match_account(description, criteria, account_name=""):
    phrases = split_criteria(criteria)

    # Context rules for common ATK / printing / BBM descriptions.
    desc = normalize_text(description)
    account = normalize_text(account_name)

    context_score = 0.0
    context_matches = []

    if re.search(r"\b(form|formulir)\b", desc):
        if any(x in account for x in ["atk", "foto copy", "fotokopi", "cetakan"]):
            context_score = 0.93
            context_matches.append("form/cetakan")

    if "post it" in desc:
        if any(x in account for x in ["atk", "perlengkapan kantor", "alat tulis"]):
            context_score = max(context_score, 0.96)
            context_matches.append("post it")

    if re.search(r"\bf4\b", desc) or "kertas f4" in desc:
        if any(x in account for x in ["atk", "perlengkapan kantor", "foto copy", "fotokopi"]):
            context_score = max(context_score, 0.96)
            context_matches.append("kertas F4")

    if re.search(r"\bbbm\b", desc) or "bahan bakar" in desc:
        if any(x in account for x in ["bbm", "transport", "bahan bakar"]):
            context_score = max(context_score, 0.96)
            context_matches.append("BBM")

    if "transport" in desc or "transportasi" in desc:
        if any(x in account for x in ["bbm", "transport", "transportasi"]):
            context_score = max(context_score, 0.96)
            context_matches.append("transport")

    # Sumbangan / kontribusi ke kegiatan sosial kemasyarakatan resmi
    # (RT/RW, desa, kelurahan, HUT RI, dsb), selama bukan ke ORMAS.
    # Dicek terhadap teks Kriteria Transaksi akun itu sendiri, bukan nama
    # akun, karena rumusan lengkapnya ada di kolom Kriteria Transaksi COA.
    criteria_norm = normalize_text(criteria)
    if (
        any(hint in desc for hint in GOV_COMMUNITY_HINTS)
        and not any(hint in desc for hint in ORMAS_HINTS)
        and any(
            phrase in criteria_norm
            for phrase in ["sosial kemasyarakatan", "pemerintahan resmi", "bencana alam"]
        )
    ):
        context_score = max(context_score, 0.93)
        context_matches.append("kegiatan sosial kemasyarakatan (pemerintahan resmi, bukan ORMAS)")

    scored = []
    for phrase in phrases:
        score = score_phrase(description, phrase)
        if score > 0:
            scored.append((phrase, score))

    scored.sort(key=lambda x: x[1], reverse=True)

    if context_score > 0:
        scored.insert(0, (context_matches[0], context_score))

    if not scored:
        return 0.0, []

    # Avoid a long list of weak matches.
    return scored[0][1], [x[0] for x in scored[:5]]


def classify_transaction(description, current_account, coa_df, uploaded_account_files=None):
    uploaded_account_files = uploaded_account_files or {}

    current_row = coa_df[coa_df["KODE AKUN"] == current_account]
    if current_row.empty:
        return {
            "current_score": 0.0, "status": "REVIEW", "matched": [],
            "alternative_code": "", "alternative_name": "",
            "alternative_file": "",
            "reason": "Kode akun tidak ditemukan di COA.",
        }

    current = current_row.iloc[0]
    current_score, current_matched = match_account(
        description,
        str(current["KRITERIA TRANSAKSI"]),
        str(current["NAMA AKUN"]),
    )

    candidates = []
    for _, row in coa_df.iterrows():
        criteria = str(row["KRITERIA TRANSAKSI"]).strip()
        if not criteria:
            continue

        score, matched = match_account(
            description,
            criteria,
            str(row["NAMA AKUN"]),
        )
        if score > 0:
            candidates.append({
                "score": score,
                "code": str(row["KODE AKUN"]),
                "name": str(row["NAMA AKUN"]),
                "matched": matched,
            })

    candidates.sort(key=lambda x: x["score"], reverse=True)
    alternatives = [x for x in candidates if x["code"] != current_account]
    best_alt = alternatives[0] if alternatives else None

    # Strong evidence on the current account.
    if current_score >= 0.78:
        status = "SESUAI"
        reason = "Deskripsi transaksi sesuai dengan kriteria akun."

        # Only overturn if another account is clearly stronger.
        if (
            best_alt
            and best_alt["score"] >= 0.90
            and best_alt["score"] > current_score + 0.12
        ):
            status = "SALAH KLASIFIKASI"
            reason = "Kecocokan dengan akun lain lebih kuat daripada akun saat ini."

    elif best_alt and best_alt["score"] >= 0.90:
        status = "SALAH KLASIFIKASI"
        reason = "Deskripsi lebih sesuai dengan kriteria akun lain pada COA."

    elif current_score >= 0.50:
        status = "REVIEW"
        reason = "Ada indikasi kecocokan, tetapi belum cukup kuat."

    else:
        status = "TIDAK SESUAI"
        reason = "Deskripsi tidak cukup cocok dengan kriteria akun."

    # Jika akun saat ini tidak punya Kriteria Transaksi sama sekali di COA,
    # skor 0 di sini bukan berarti transaksi salah, tapi data COA yang
    # belum lengkap. Tandai secara eksplisit supaya tidak tercampur
    # dengan transaksi yang memang benar-benar tidak sesuai.
    if not str(current["KRITERIA TRANSAKSI"]).strip():
        status = "REVIEW"
        reason = "Kriteria Transaksi untuk akun ini masih kosong di COA."

    alt_code = best_alt["code"] if best_alt else ""
    alt_name = (
        f'{best_alt["code"]} - {best_alt["name"]}'
        if best_alt else ""
    )
    alt_file = "; ".join(uploaded_account_files.get(alt_code, []))

    return {
        "current_score": current_score,
        "status": status,
        "matched": current_matched,
        "alternative_code": alt_code,
        "alternative_name": alt_name,
        "alternative_file": alt_file,
        "reason": reason,
    }


def analyze_transactions(tx_df, coa_df):
    uploaded_account_files = {}
    for _, row in tx_df[["KODE AKUN", "SOURCE FILE"]].drop_duplicates().iterrows():
        uploaded_account_files.setdefault(
            str(row["KODE AKUN"]), []
        ).append(str(row["SOURCE FILE"]))

    results = []

    for _, row in tx_df.iterrows():
        c = classify_transaction(
            row["DESCRIPTION"],
            row["KODE AKUN"],
            coa_df,
            uploaded_account_files,
        )

        criteria_rows = coa_df[coa_df["KODE AKUN"] == row["KODE AKUN"]]
        criteria = (
            str(criteria_rows.iloc[0]["KRITERIA TRANSAKSI"])
            if not criteria_rows.empty else ""
        )

        results.append({
            "SOURCE FILE": row["SOURCE FILE"],
            "KODE AKUN": row["KODE AKUN"],
            "NAMA AKUN": row["NAMA AKUN"],
            "VOUCHER NO.": row["VOUCHER NO."],
            "TRANS. DATE": row["TRANS. DATE"],
            "ENTRY DATE": row["ENTRY DATE"],
            "DESCRIPTION": row["DESCRIPTION"],
            "DEBIT BASE": row["DEBIT BASE"],
            "CREDIT BASE": row["CREDIT BASE"],
            "DEBIT FOREX": row["DEBIT FOREX"],
            "CREDIT FOREX": row["CREDIT FOREX"],
            "DOCUMENT NO.": row["DOCUMENT NO."],
            "COA CRITERIA": criteria,
            "STATUS": c["status"],
            "MATCH SCORE": round(c["current_score"] * 100, 1),
            "MATCHED CRITERIA": ", ".join(c["matched"]),
            "ALTERNATIVE COA": c["alternative_code"],
            "ALTERNATIVE ACCOUNT": c["alternative_name"],
            "ALTERNATIVE FILE": c["alternative_file"],
            "REVIEW REASON": c["reason"],
        })

    return pd.DataFrame(results, columns=RESULT_COLUMNS)


# ============================================================
# EXCEL OUTPUT
# ============================================================

def safe_sheet_name(name, used):
    name = re.sub(r"[\[\]\*:/\\?]", "_", str(name))
    name = name[:31] or "ACCOUNT"

    base = name
    counter = 1
    while name in used:
        suffix = f"_{counter}"
        name = (base[:31 - len(suffix)] + suffix)
        counter += 1

    used.add(name)
    return name


def style_sheet(ws, widths=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E1F2")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width


def build_excel(results_df, coa_df):
    """
    Builds a single workbook:
      - Sheet "All Data": every transaction from the uploaded file(s),
        tagged with KODE AKUN / NAMA AKUN so accounts stay identifiable.
      - One sheet per Account Name (NAMA AKUN): every transaction that
        belongs to that account, regardless of COA-matching status.

    Every sheet uses the same simple layout:
        VOUCHER NO. | TRANS. DATE | ENTRY DATE | DESCRIPTION | DEBIT | CREDIT
    DEBIT / CREDIT are taken only from the "Amount in Base CCY" columns
    (DEBIT BASE / CREDIT BASE) — forex amounts are not included.
    """
    output = io.BytesIO()
    used_sheets = set()

    detail_cols = [
        "VOUCHER NO.", "TRANS. DATE", "ENTRY DATE", "DESCRIPTION",
        "DEBIT", "CREDIT",
    ]

    # The raw MDIS export bakes branch/regional text into the account name
    # (e.g. "ATK, Foto Copy dan Cetakan - Cabang Cipanas - Regional J"),
    # which is both long and inconsistent across files. Prefer the clean
    # Nama Akun from the COA (keyed by Kode Akun) for display/sheet naming,
    # and only fall back to the raw parsed name if the code isn't in COA.
    coa_name_lookup = {}
    if coa_df is not None:
        for _, coa_row in coa_df.iterrows():
            code = str(coa_row["KODE AKUN"]).strip()
            name = str(coa_row["NAMA AKUN"]).strip()
            if code and name:
                coa_name_lookup[code] = name

    data = results_df.copy()
    data["DEBIT"] = data["DEBIT BASE"]
    data["CREDIT"] = data["CREDIT BASE"]
    data["NAMA AKUN"] = data.apply(
        lambda r: coa_name_lookup.get(str(r["KODE AKUN"]).strip(), r["NAMA AKUN"]),
        axis=1,
    )

    # Sort chronologically within each account without changing the
    # original text formatting of TRANS. DATE.
    data["_SORT_DATE"] = pd.to_datetime(
        data["TRANS. DATE"], dayfirst=True, errors="coerce"
    )
    data = data.sort_values(["KODE AKUN", "_SORT_DATE"]).drop(columns=["_SORT_DATE"])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ---- "All Data" sheet: everything, with account identifiers ----
        all_data_cols = ["KODE AKUN", "NAMA AKUN"] + detail_cols
        all_data_sheet = safe_sheet_name("All Data", used_sheets)
        data[all_data_cols].to_excel(writer, sheet_name=all_data_sheet, index=False)

        ws = writer.book[all_data_sheet]
        style_sheet(ws, {
            "A": 13, "B": 40, "C": 24, "D": 13, "E": 13,
            "F": 60, "G": 18, "H": 18,
        })

        # ---- One sheet per Account Name, in COA order when possible ----
        coa_order = list(coa_df["KODE AKUN"]) if coa_df is not None else []
        codes_present = [c for c in data["KODE AKUN"].unique().tolist() if c]
        ordered_codes = (
            [c for c in coa_order if c in codes_present]
            + [c for c in codes_present if c not in coa_order]
        )

        for code in ordered_codes:
            group = data[data["KODE AKUN"] == code]
            names = group["NAMA AKUN"].dropna()
            names = names[names.astype(str).str.strip() != ""]
            account_name = names.iloc[0] if not names.empty else code

            sheet_title = safe_sheet_name(account_name, used_sheets)
            group[detail_cols].to_excel(writer, sheet_name=sheet_title, index=False)

            ws = writer.book[sheet_title]
            style_sheet(ws, {
                "A": 24, "B": 13, "C": 13, "D": 60, "E": 18, "F": 18,
            })

    output.seek(0)
    return output.getvalue()


# ============================================================
# STREAMLIT UI
# ============================================================

st.set_page_config(
    page_title="COA Transaction Checker",
    layout="wide",
)

st.title("COA Transaction Checker")
st.caption(
    "Membandingkan setiap transaksi dengan Kriteria Transaksi akun saat ini "
    "dan seluruh akun pada COA untuk mendeteksi salah klasifikasi."
)

with st.expander("Cara download file `.xls` dan daftar akun yang dicakup", expanded=False):
    st.markdown("""
**Cara download 1 file `.xls` yang berisi transaksi seluruh akun:**

> **Catatan:** Cukup download **1 file `.xls` untuk seluruh akun** sesuai range akun yang ditentukan. **Tidak perlu download file satu per satu per akun dan tidak perlu menggabungkan beberapa file secara manual.**

1. Buka menu **Transaction by account**
2. **Transaction date** diisi tanggal yang akan dicek (biasanya periode audit)
3. **From Account No.:** `54100000-KODE CABANG-KODE REGIONAL` *(Contoh: 54100000-024-10)*
4. **To Account No.:** `59924000-KODE CABANG-KODE REGIONAL` *(Contoh: 59924000-024-10)*
5. **Source Code** diisi: `TPB-KODE CABANG` *(Contoh: TPB-024)*
6. **Update status:** All, **Type:** Printing
7. Klik **Kirim ke Excel**, kemudian simpan file `.xls` yang dihasilkan.
8. File `.xls` yang sudah disimpan **langsung di-upload ke menu Upload file transaksi**.
""")
    st.image("panduan.png", use_container_width=True)
    st.divider()

    st.markdown("**Daftar 24 akun yang dicakup:**")
    akun_df = pd.DataFrame(
        [{"KODE AKUN": k, "NAMA AKUN": v} for k, v in ALLOWED_ACCOUNTS.items()]
    )
    st.dataframe(akun_df, use_container_width=True, hide_index=True)


with st.expander("Disclaimer", expanded=False):
    st.warning(
        "Hasil klasifikasi di tool ini bersifat indikatif, "
        "bukan kesimpulan final. Deskripsi transaksi dari MDIS sangat bebas "
        "formatnya (typo, singkatan tidak konsisten, urutan kata berubah-ubah), "
        "sehingga pencocokan otomatis tidak bisa menangkap semua variasi. "
        "Status REVIEW, TIDAK SESUAI, dan SALAH KLASIFIKASI tetap perlu "
        "diverifikasi manual oleh auditor sebelum dijadikan dasar keputusan."
    )

st.divider()
st.subheader("1. Sumber Data")

with st.container(border=True):
    input_col1, input_col2 = st.columns([3, 2], gap="large")

    with input_col1:
        coa_url = st.text_input(
            "Google Sheets URL (COA)",
            value=DEFAULT_COA_URL,
            help="Gunakan link view-only Google Sheets yang berisi Kode Akun, Nama Akun, dan Kriteria Transaksi.",
        )
        refresh_col, _ = st.columns([1, 4])
        with refresh_col:
            if st.button("Refresh COA", use_container_width=True):
                load_coa_from_google.clear()
                st.rerun()

    with input_col2:
        single_upload = st.file_uploader(
            "Upload file transaksi (.xls)",
            type=["xls"],
            accept_multiple_files=False,
            help="Cukup 1 file export MDIS yang sudah berisi seluruh akun.",
        )
        uploaded_files = [single_upload] if single_upload is not None else []

    st.caption(
        "Threshold: SESUAI >= 78%. REVIEW 50-77.9%. "
        "TIDAK SESUAI < 50%. Akun alternatif hanya dianggap salah klasifikasi "
        "jika kecocokannya kuat dan jelas lebih tinggi. Akun dengan Kriteria "
        "Transaksi kosong di COA otomatis ditandai REVIEW, bukan TIDAK SESUAI."
    )

if not coa_url:
    st.warning("Masukkan URL Google Sheets COA.")
    st.stop()

try:
    coa_df = load_coa_from_google(coa_url)
except Exception as exc:
    st.error(f"Gagal membaca COA: {exc}")
    st.info(
        "Pastikan Google Sheets dapat dibuka tanpa login "
        "dan minimal kolom A-C berisi Kode Akun, Nama Akun, dan Kriteria Transaksi."
    )
    st.stop()

if coa_df.empty:
    st.error("COA tidak memiliki data yang dapat digunakan.")
    st.stop()

st.divider()
st.subheader("2. Ringkasan COA")

col1, col2, col3 = st.columns(3)
col1.metric("Jumlah akun COA", len(coa_df))
col2.metric(
    "Akun biaya",
    int(
        coa_df["KELOMPOK"]
        .astype(str)
        .str.upper()
        .str.contains("BIAYA", na=False)
        .sum()
    ),
)
col3.metric(
    "Kriteria terisi",
    int((coa_df["KRITERIA TRANSAKSI"].astype(str).str.strip() != "").sum()),
)

if not uploaded_files:
    st.info("Upload file Transaction Listing by Accounts (1 file, semua akun) untuk mulai analisis.")
    st.stop()

all_transactions = []
parse_errors = []
no_data_files = []

for uploaded_file in uploaded_files:
    try:
        parsed = parse_uploaded_xls(uploaded_file)

        if parsed.empty:
            no_data_files.append(uploaded_file.name)
            continue

        all_transactions.append(parsed)
    except Exception as exc:
        parse_errors.append(f"{uploaded_file.name}: {exc}")

if no_data_files:
    st.info(
        "File ini valid tetapi 0 transaksi: "
        + ", ".join(no_data_files)
    )

if parse_errors:
    st.warning("Ada file yang benar-benar gagal dibaca:")
    for error in parse_errors:
        st.write(f"- {error}")

if not all_transactions:
    st.error("Tidak ada file yang berhasil diproses.")
    st.stop()

transactions = pd.concat(all_transactions, ignore_index=True)

out_of_scope = transactions[~transactions["KODE AKUN"].isin(ALLOWED_ACCOUNTS)]
if not out_of_scope.empty:
    out_of_scope_summary = (
        out_of_scope[["KODE AKUN", "NAMA AKUN", "SOURCE FILE"]]
        .drop_duplicates()
        .reset_index(drop=True)
    )
    st.warning(
        f"{len(out_of_scope_summary)} kombinasi akun/file di luar 24 akun "
        "yang dicakup tool ini terdeteksi. Transaksinya tetap diproses, "
        "tapi mohon dicek apakah file yang diupload sudah benar."
    )
    st.dataframe(out_of_scope_summary, use_container_width=True, hide_index=True)

results = analyze_transactions(transactions, coa_df)

st.divider()
st.subheader("3. Hasil Analisis")

metric_cols = st.columns(5)
metric_cols[0].metric("Total transaksi", len(results))
metric_cols[1].metric("Sesuai", int((results["STATUS"] == "SESUAI").sum()))
metric_cols[2].metric("Review", int((results["STATUS"] == "REVIEW").sum()))
metric_cols[3].metric("Tidak sesuai", int((results["STATUS"] == "TIDAK SESUAI").sum()))
metric_cols[4].metric(
    "Salah klasifikasi",
    int((results["STATUS"] == "SALAH KLASIFIKASI").sum()),
)

filter_col1, filter_col2, filter_col3 = st.columns(3)

with filter_col1:
    status_filter = st.multiselect(
        "Filter status",
        options=sorted(results["STATUS"].unique()),
        default=sorted(results["STATUS"].unique()),
    )

with filter_col2:
    account_filter = st.multiselect(
        "Filter akun",
        options=sorted(results["KODE AKUN"].unique()),
        default=sorted(results["KODE AKUN"].unique()),
    )

with filter_col3:
    min_score = st.slider(
        "Minimum match score",
        min_value=0,
        max_value=100,
        value=0,
        step=5,
    )

filtered = results[
    results["STATUS"].isin(status_filter)
    & results["KODE AKUN"].isin(account_filter)
    & (results["MATCH SCORE"] >= min_score)
].copy()

# Download is intentionally placed BEFORE the large result tables so the
# user never needs to scroll to the bottom or zoom out.
excel_bytes = build_excel(results, coa_df)

dl1, dl2 = st.columns([1, 5])
with dl1:
    st.download_button(
        label="📥 Download XLSX",
        data=excel_bytes,
        file_name="rekap_transaksi_per_akun.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

st.dataframe(
    filtered,
    use_container_width=True,
    height=360,
    hide_index=True,
    column_config={
        "MATCH SCORE": st.column_config.ProgressColumn(
            "MATCH SCORE",
            min_value=0,
            max_value=100,
            format="%.1f%%",
        ),
        "DEBIT BASE": st.column_config.NumberColumn(
            "DEBIT BASE",
            format="%,.0f",
        ),
        "CREDIT BASE": st.column_config.NumberColumn(
            "CREDIT BASE",
            format="%,.0f",
        ),
        "DEBIT FOREX": st.column_config.NumberColumn(
            "DEBIT FOREX",
            format="%,.0f",
        ),
        "CREDIT FOREX": st.column_config.NumberColumn(
            "CREDIT FOREX",
            format="%,0.00",
        ),
    },
)

# Highlight potential misclassification in a compact review table.
review = results[
    results["STATUS"].isin(["TIDAK SESUAI", "SALAH KLASIFIKASI", "REVIEW"])
].copy()

if not review.empty:
    st.divider()
    st.subheader("🔎 Prioritas Review Auditor")
    st.dataframe(
        review[
            [
                "KODE AKUN",
                "NAMA AKUN",
                "VOUCHER NO.",
                "DESCRIPTION",
                "STATUS",
                "MATCH SCORE",
                "ALTERNATIVE COA",
                "ALTERNATIVE ACCOUNT",
                "ALTERNATIVE FILE",
                "REVIEW REASON",
            ]
        ],
        use_container_width=True,
        height=360,
        hide_index=True,
    )

st.caption(
    "Output: 1 workbook, sheet 'All Data' berisi seluruh transaksi, "
    "ditambah 1 sheet per Account Name (Nama Akun) berisi transaksi akun "
    "tersebut. Kolom DEBIT/CREDIT pada file Excel diambil dari Amount in "
    "Base CCY (bukan forex)."
)
